import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
from PIL import Image, ImageTk
from tkinter import ttk

# ===============================
#  BACK-END
# ===============================

def quebrar_celulas_mescladas(ws):
    """
    Desmescla todas as células de uma planilha, copiando o valor da
    célula superior esquerda para todas as células que faziam parte
    do intervalo mesclado.
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        valor = ws.cell(row=min_row, column=min_col).value

        # Desmescla o intervalo
        ws.unmerge_cells(range_string=str(merged_range))

        # Preenche as células do intervalo com o valor original
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=valor)
    return ws

def carregar_planilha_mestre_auto(self):
    diretorio_mestre = "./diretorio_mestre/"
    try:
        arquivos = [arq for arq in os.listdir(diretorio_mestre) if arq.lower().endswith(('.xls', '.xlsx', '.xlsb'))]
        if not arquivos:
            raise FileNotFoundError("Nenhum arquivo Excel encontrado no diretório mestre.")
        
        caminho_completo = os.path.join(diretorio_mestre, arquivos[0])
        self.caminho_mestre = caminho_completo
        self.label_mestre.config(text=f"Mestre: {arquivos[0]}")

    except Exception as e:
        self.label_mestre.config(text=f"Erro: {e}")
        messagebox.showerror("Erro", str(e))


def encontrar_cabecalho_personalizado(df, colunas_busca, max_linhas=15):
    """
    Tenta encontrar, nas primeiras 'max_linhas' do DataFrame,
    as colunas especificadas em 'colunas_busca', mesmo que
    estejam em linhas diferentes.

    Retorna:
    - Uma lista de nomes de colunas unificados, se encontrado
    - O índice (linha) até onde foi usado para compor o cabeçalho
    """
    num_cols = df.shape[1]
    header_acumulado = [""] * num_cols
    linha_final_cabecalho = 0

    limite = min(max_linhas, len(df))
    for i in range(limite):
        row_values = df.iloc[i].fillna("").astype(str).tolist()
        for c in range(num_cols):
            val = row_values[c].strip()
            if val:
                if header_acumulado[c] == "":
                    header_acumulado[c] = val
                else:
                    header_acumulado[c] += f" {val}"
        colunas_encontradas = [h for h in header_acumulado if any(b in h for b in colunas_busca)]
        if len(colunas_encontradas) >= len(colunas_busca):
            linha_final_cabecalho = i
            break

    return header_acumulado, linha_final_cabecalho

def carregar_planilha_e_filtrar(caminho_arquivo):
    """
    Carrega a planilha, desmescla as células (apenas se não for .xlsb)
    e filtra os dados com base nas colunas necessárias e nos valores
    'SIM' ou 'X' nas colunas 'Optante de transporte' e 'Usará transporte na HE'.
    Agora a comparação é por 'Reg.' (renomeado para 'Registro').
    """
    # Se for .xlsb, carregamos diretamente com pandas (pyxlsb)
    # e PULAMOS o trecho de openpyxl (pois não há suporte para .xlsb).
    if caminho_arquivo.lower().endswith(".xlsb"):
        df = pd.read_excel(
            caminho_arquivo,
            engine='pyxlsb',
            header=None
        )
    else:
        df = pd.read_excel(
            caminho_arquivo,
            engine=None,  # openpyxl por padrão
            header=None
        )
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        ws = quebrar_celulas_mescladas(ws)
        wb.save(caminho_arquivo)

        # Recarrega após desmesclar
        df = pd.read_excel(
            caminho_arquivo,
            engine=None,
            header=None
        )

    # Colunas que queremos identificar no cabeçalho
    colunas_desejadas = [
        "Reg.", "Nome empregado", "Unidade de Negócio", "Turno",
        "Optante de transporte", "Usará transporte na HE", "LANCHE",
        "HORARIO DE SAÍDA", "OBSERVAÇÃO"
    ]

    # Monta o cabeçalho personalizadamente
    header_unificado, idx_cabecalho = encontrar_cabecalho_personalizado(df, colunas_desejadas, max_linhas=15)
    df.columns = header_unificado
    df = df.iloc[idx_cabecalho + 1:].copy()

    # Mantém só as colunas de interesse
    colunas_existentes = [col for col in df.columns if any(c.lower() in col.lower() for c in colunas_desejadas)]
    if not colunas_existentes:
        raise ValueError("Nenhuma das colunas necessárias foi encontrada na planilha.")

    df = df[colunas_existentes].copy()

    # Renomeia colunas para o padrão que usaremos
    for c in df.columns:
        if "reg." in c.lower():
            df.rename(columns={c: "Registro"}, inplace=True)
        if "optante de transporte" in c.lower():
            df.rename(columns={c: "Optante de transporte"}, inplace=True)
        if "usará transporte na he" in c.lower():
            df.rename(columns={c: "Usará transporte na HE"}, inplace=True)
        if "nome empregado" in c.lower():
            df.rename(columns={c: "Nome empregado"}, inplace=True)

    # Remove linhas onde "Nome empregado" é NaN (opcional)
    if "Nome empregado" in df.columns:
        df = df[df["Nome empregado"].notna()]

    # Normaliza valores de transporte (SIM ou X)
    if "Optante de transporte" in df.columns:
        df["Optante de transporte"] = (
            df["Optante de transporte"].astype(str).str.upper().str.strip()
        )
    if "Usará transporte na HE" in df.columns:
        df["Usará transporte na HE"] = (
            df["Usará transporte na HE"].astype(str).str.upper().str.strip()
        )

    # Filtra somente quem tiver SIM ou X em ambas as colunas
    if "Optante de transporte" in df.columns and "Usará transporte na HE" in df.columns:
        df = df[
            (df["Optante de transporte"].isin(["SIM", "X"])) &
            (df["Usará transporte na HE"].isin(["SIM", "X"]))
        ]

    return df

def comparar_planilhas(caminho_mestre, caminhos_comparacao):
    """
    Compara a planilha mestre com diversas planilhas de comparação
    pela coluna 'Registro' (removendo zeros à esquerda para ambas).
    """
    if not caminho_mestre or not caminhos_comparacao:
        raise ValueError("Selecione a planilha mestre e as planilhas para comparação.")

    # Carrega a planilha mestre
    try:
        planilha_mestre = pd.read_excel(caminho_mestre, header=None)
        planilha_mestre.columns = [
            "Linha", "Turno", "Itinerário", "Registro",
            "Nome dos Passageiros", "Endereço", "Bairro", "Telefone"
        ]
        # Remove zeros à esquerda na planilha mestre
        planilha_mestre["Registro"] = (
            planilha_mestre["Registro"]
            .astype(str)
            .str.strip()
            .str.lstrip("0")  # <-- REMOVE zeros à esquerda
        )
    except Exception as e:
        raise ValueError(f"Erro ao processar a planilha mestre: {e}")

    # Carrega e filtra as planilhas de comparação
    dfs = []
    for caminho in caminhos_comparacao:
        try:
            df_filtrado = carregar_planilha_e_filtrar(caminho)
            if df_filtrado.empty:
                print(f"A planilha {os.path.basename(caminho)} não contém dados válidos após filtragem.")
                continue
            if "Registro" not in df_filtrado.columns:
                print(f"A planilha {os.path.basename(caminho)} não possui coluna 'Reg.' / 'Registro'.")
                continue

            # Remove zeros à esquerda nas planilhas de comparação
            df_filtrado["Registro"] = (
                df_filtrado["Registro"]
                .astype(str)
                .str.strip()
                .str.lstrip("0")  # <-- REMOVE zeros à esquerda
            )

            dfs.append(df_filtrado)
        except ValueError as e:
            print(f"Erro ao filtrar a planilha {os.path.basename(caminho)}: {e}")
            continue
        except Exception as e:
            print(f"Erro inesperado ao carregar a planilha {os.path.basename(caminho)}: {e}")
            continue

    if not dfs:
        raise ValueError("Nenhuma planilha válida foi encontrada para comparação.")

    # Concatena todas as planilhas de comparação
    try:
        todas_planilhas = pd.concat(dfs, ignore_index=True, axis=0)
    except ValueError as e:
        raise ValueError("Erro ao concatenar as planilhas: " + str(e))

    # Remove linhas completamente vazias
    todas_planilhas = todas_planilhas.dropna(how='all')

    # Compara somente por 'Registro' (já sem zeros à esquerda)
    comparacao = planilha_mestre[
        planilha_mestre["Registro"].isin(todas_planilhas["Registro"])
    ]

    # Ordena por Turno e Itinerário
    comparacao["Itinerário"] = comparacao["Itinerário"].astype(str)
    comparacao = comparacao.sort_values(by=["Turno", "Itinerário"])

    # Cria o layout final, separando por turno e itinerário
    separada_por_turnos = []
    for turno, grupo_turno in comparacao.groupby("Turno"):
        # Linha separadora do turno
        separator_turno = pd.DataFrame([[""] * len(comparacao.columns)], columns=comparacao.columns)
        separator_turno.at[0, "Turno"] = f"Turno: {turno}"
        separada_por_turnos.append(separator_turno)

        # Linha vazia (opcional)
        linha_vazia = pd.DataFrame([[""] * len(comparacao.columns)], columns=comparacao.columns)
        separada_por_turnos.append(linha_vazia)

        # Para cada itinerário dentro do turno
        for itinerario, grupo_itinerario in grupo_turno.groupby("Itinerário"):
            # Cabeçalho
            cabecalho_df = pd.DataFrame([planilha_mestre.columns], columns=planilha_mestre.columns)
            separada_por_turnos.append(cabecalho_df)

            # Dados do itinerário
            separada_por_turnos.append(grupo_itinerario)

            # Duas linhas vazias
            linhas_vazias = pd.DataFrame([[""] * len(grupo_itinerario.columns)] * 2, columns=grupo_itinerario.columns)
            separada_por_turnos.append(linhas_vazias)

    planilha_final = pd.concat(separada_por_turnos, ignore_index=True)
    return planilha_final

def gerar_nome_sheet_com_data():
    """
    Gera o nome da aba (sheet) no formato dd.mm,
    por exemplo: '10.03'.
    """
    data_atual = datetime.now()
    return data_atual.strftime("%d.%m")

def salvar_planilha_com_estilo(planilha, caminho_saida):
    """
    Salva o DataFrame 'planilha' em um arquivo Excel,
    aplicando estilos e formatação com openpyxl.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    nome_sheet = gerar_nome_sheet_com_data()
    ws.title = nome_sheet

    estilo_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    estilo_azul_claro = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    estilo_laranja = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    estilo_verde_claro = PatternFill(start_color="0cff00", end_color="0cff00", fill_type="solid")  # Novo estilo verde claro
    estilo_negrito = Font(bold=True)
    centralizado = Alignment(horizontal="center", vertical="center")

    border_style = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000")
    )

    # Cabeçalho (primeira linha)
    for col_num, col_name in enumerate(planilha.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = estilo_negrito
        cell.alignment = centralizado
        if col_name in ["Linha", "Turno", "Itinerário", "Registro"]:
            cell.fill = estilo_amarelo
        else:
            cell.fill = estilo_azul_claro
        cell.border = border_style

    # Demais linhas
    for row_num, row in enumerate(planilha.itertuples(index=False, name=None), start=2):
        linha_vazia = all(value == "" for value in row)
        linha_turno = any(isinstance(value, str) and value.startswith("Turno:") for value in row)

        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)

            # Se for linha de turno, aplica o estilo laranja
            if linha_turno:
                cell.fill = estilo_laranja
                cell.font = estilo_negrito
                cell.alignment = centralizado
            else:
                # Se for linha de cabeçalho repetida
                if str(value).upper().strip() in [str(col).upper().strip() for col in planilha.columns]:
                    cell.font = estilo_negrito
                    # Preenchimento de acordo com a coluna
                    if planilha.columns[col_num - 1] in ["Linha", "Turno", "Itinerário", "Registro"]:
                        cell.fill = estilo_amarelo
                    else:
                        cell.fill = estilo_azul_claro
                    cell.alignment = centralizado
                else:
                    cell.alignment = centralizado

            # Aplica borda somente se a linha não for completamente vazia
            if not linha_vazia:
                cell.border = border_style

            # Verifica se a coluna é "Bairro" e se o valor termina com "Jac." ou "Jacareí"
            if planilha.columns[col_num - 1] == "Bairro" and isinstance(value, str):
                if value.strip().endswith("Jac.") or value.strip().endswith("Jacareí"):
                    cell.fill = estilo_verde_claro

    wb.save(caminho_saida)

def gerar_nome_arquivo_sugerido():
    pasta_destino = r"C:/Comparador de Planilhas/Viagens do dia/"
    data_hoje = datetime.now().strftime("%Y-%m-%d")

    # Lista arquivos existentes para essa data
    arquivos_existentes = [f for f in os.listdir(pasta_destino) if f.startswith(data_hoje) and f.endswith(".xlsx")]

    numero = 1
    while True:
        sufixo = f".{numero:02d}"
        nome_arquivo = f"{data_hoje}{sufixo}.xlsx"
        if nome_arquivo not in arquivos_existentes:
            break
        numero += 1

    caminho_final = os.path.join(pasta_destino, nome_arquivo)
    return caminho_final

# ===============================
#  FRONT-END (Tkinter)
# ===============================

class ComparadorPlanilhas:
    def __init__(self, root):
        style = ttk.Style()
        self.root = root
        self.root.title("Comparador de Planilhas Excel")
        self.root.geometry("1200x700")
        self.root.config(bg="#0a0f2c")

        self.caminho_mestre = None
        self.caminhos_comparacao = []


        # Container com azul claro e "opacidade"
        # Container com azul claro e borda verde
        self.container = tk.Frame(
            root, 
            bg="#4e6ca8",
            highlightbackground="#e0f0ff",  
            highlightcolor="#e0f0ff",       
            highlightthickness=3,           # Espessura da borda
            bd=0
        )
        self.container.place(relx=0.5, rely=0.5, anchor="center", width=450, height=550)

        # Agora um Frame INTERNO para os botões dentro do container
        self.inner_frame = tk.Frame(self.container, bg="#4e6ca8")
        self.inner_frame.place(relx=0.5, rely=0.5, anchor="center")






        # Título

        # Título dentro do container azul claro
        titulo = tk.Label(
            self.inner_frame, 
            text=" COMPARADOR DE PLANILHAS ", 
            font=("Helvetica", 20, "bold"),
            bg="#4e6ca8", 
            fg="white"
        )
        titulo.pack(pady=(55, 20))






        linha = tk.Canvas(root, height=2, bg="white", bd=0, relief="sunken")
        linha.pack(fill="x", padx=10, pady=5)

        self.container = tk.Frame(root, bg="#282c34")
        self.container.pack(pady=10)



        

        # Botão Planilha Mestre
        self.botao_mestre = tk.Button(
            self.inner_frame,
            text="Selecionar Mestre",
            command=self.selecionar_mestre,
            bg="#1c294a",
            fg="#ffffff",
            font=("Arial", 12, "bold"),
            relief="flat",
            bd=0,
            activebackground="#263b6a"
        )
        self.botao_mestre.pack(pady=5, ipadx=10, ipady=5)
        self._aplicar_estilo_botao(self.botao_mestre)

        # Label Mestre
        self.label_mestre = tk.Label(
            self.inner_frame,
            text="Mestre não selecionada.",
            bg="#4e6ca8",
            fg="#ffffff",
            font=("Arial", 10,"bold")
        )
        self.label_mestre.pack(pady=(0, 10))



        # Botão Carregar Planilhas
        self.botao_comparacao = tk.Button(
            self.inner_frame,
            text="Recarregar Planilhas",
            command=self.carregar_comparacao_automatica,
            bg="#1c294a",
            fg="#ffffff",
            font=("Arial", 12, "bold"),
            relief="flat",
            bd=0,
            activebackground="#263b6a"
        )
        self.botao_comparacao.pack(pady=5, ipadx=10, ipady=5)
        self._aplicar_estilo_botao(self.botao_comparacao)

        

        # Label Comparação (fixa, sem recriar depois)
        self.label_comparacao = tk.Label(
            self.inner_frame,
            text="Nenhuma planilha carregada.",
            bg="#005a92",
            fg="white",
            font=("Arial", 10,"bold"),
            width=30
        )
        self.label_comparacao.pack(pady=(0, 15))

       

        # Frame horizontal para os dois botões
        self.frame_horizontal = tk.Frame(self.inner_frame, bg="#4e6ca8")
        self.frame_horizontal.pack(pady=5)



       # Botão Comparar Planilhas
        self.botao_comparar = tk.Button(
            self.frame_horizontal,
            text="Comparar Planilhas",
            command=self.comparar_planilhas,
            bg="#920000",
            fg="#ffffff",
            font=("Arial", 12, "bold"),
            relief="flat",
            bd=0,
            activebackground="#263b6a",
            width=15  # <-- controla o tamanho
        )
        self.botao_comparar.pack(side="left", padx=5, ipadx=5, ipady=5)
        self._aplicar_estilo_botao(self.botao_comparar)


        self.label_movido = tk.Label(
            self.inner_frame,
            text="",  # Vazio inicialmente
            bg="#005a92",
            fg="#00ff00",
            font=("Arial", 10, "bold")
        )
        self.label_movido.pack(pady=(5, 0))  # Espaçamento leve abaixo do botão



        # Botão Abrir Pasta
        self.botao_abrir_pasta = tk.Button(
            self.frame_horizontal,
            text="Abrir Pasta",
            command=self.abrir_pasta_viagens,
            bg="#1c294a",
            fg="#ffffff",
            font=("Arial", 12, "bold"),
            relief="flat",
            bd=0,
            activebackground="#263b6a",
            width=15  # <-- mesmo tamanho do anterior
        )
        self.botao_abrir_pasta.pack(side="left", padx=5, ipadx=5, ipady=5)
        self._aplicar_estilo_botao(self.botao_abrir_pasta)


        # Botão Sair
        self.botao_sair = tk.Button(
            self.inner_frame,
            text="Sair",
            command=self.root.quit,
            bg="#1c294a",
            fg="#ffffff",
            font=("Arial", 12, "bold"),
            relief="flat",
            bd=0,
            activebackground="#263b6a"
        )
        self.botao_sair.pack(pady=(25, 10), ipadx=10, ipady=5)
        self._aplicar_estilo_botao(self.botao_sair)

        try:
            imagem_logo = Image.open("LOGO.jpeg")
            imagem_logo = imagem_logo.resize((175, 70))
            self.logo_img = ImageTk.PhotoImage(imagem_logo)

            self.label_logo = tk.Label(root, image=self.logo_img, bg="#282c34")
            self.label_logo.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-10)

            self.label_assinatura = tk.Label(
                root,
                text="Designed by HR IT Services © 2025",
                bg="#0a0f2c",
                fg="white",
                font=("Arial", 6, "italic")
            )
            self.label_assinatura.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-85)
            
        except Exception as e:
            print(f"Erro ao carregar LOGO.jpeg: {e}")





    
    def _aplicar_estilo_botao(self, botao):
        botao.configure(
            cursor="hand2",
            bg="#1c294a",              # cor do botão (background escuro)
            fg="white",                # texto branco
            font=("Arial", 12, "bold"),
            bd=3,                      # borda fina
            relief="ridge",            # dá um leve "depth"
            highlightthickness=1,
            highlightbackground="white"  # borda branca fina
        )

        # Estado personalizado para controle do Hover
        botao.hover_ativo = True
        botao.after_id = None  # ID do temporizador do botão

        def on_enter(e):
            if botao.hover_ativo:
                botao.config(bg="#263b6a")

        def on_leave(e):
            if botao.hover_ativo:
                botao.config(bg="#1c294a")

        botao.bind("<Enter>", on_enter)
        botao.bind("<Leave>", on_leave)


        # Hover com efeito de "sombra" mais clara
        botao.bind("<Enter>", lambda e: botao.config(
            bg="#263b6a",
            relief="raised",
            bd=2,
            highlightthickness=1
        ))
        botao.bind("<Leave>", lambda e: botao.config(
            bg="#1c294a",
            relief="ridge",
            bd=1,
            highlightthickness=1
        ))

        
        botao.bind("<Enter>", lambda e: botao.config(bg="#263b6a"))
        botao.bind("<Leave>", lambda e: botao.config(bg="#1c294a"))



    def selecionar_mestre(self):
        diretorio_mestre = r"C:/Comparador de Planilhas/Masterdata/"
        diretorio_backup = r"C:/Comparador de Planilhas/Backup/"
        try:
            arquivos = [arq for arq in os.listdir(diretorio_mestre) if arq.lower().endswith(('.xls', '.xlsx', '.xlsb'))]
            if not arquivos:
                raise FileNotFoundError("Nenhum arquivo Excel encontrado no diretório mestre.")
    
            self.caminho_mestre = os.path.join(diretorio_mestre, arquivos[0])
            self.label_mestre.config(text=f"Mestre: {arquivos[0]}")
        
            # Geração de backup automático
            os.makedirs(diretorio_backup, exist_ok=True)
            data_hoje = datetime.now().strftime("%Y-%m-%d")
            base_nome = f"Planilha Mestre {data_hoje}"
            numero = 1

            while True:
                sufixo = f".{numero:02d}"
                nome_backup = f"{base_nome}{sufixo}{os.path.splitext(self.caminho_mestre)[1]}"
                caminho_backup = os.path.join(diretorio_backup, nome_backup)
                if not os.path.exists(caminho_backup):
                    break
                numero += 1

            import shutil
            shutil.copy2(self.caminho_mestre, caminho_backup)

            # Indicação visual de sucesso
            self.botao_mestre.config(bg="#4CAF50")  # Verde
            self.label_mestre.config(text=f"Mestre: {arquivos[0]} \n BACKUP CRIADO!")

            self.root.after(5000, lambda: self.reset_cor_botao(self.botao_mestre))

        except Exception as e:
            self.botao_mestre.config(bg="#F44336")  # Vermelho
            self.label_mestre.config(text=f"Erro: {e}")

            self.root.after(3000, lambda: self.reset_cor_botao(self.botao_mestre))


    def carregar_comparacao_automatica(self):
        diretorio_comparacao = r"C:/Comparador de Planilhas/Extras Planilhas/"
        try:
            arquivos = [os.path.join(diretorio_comparacao, arq) 
                        for arq in os.listdir(diretorio_comparacao) 
                        if arq.lower().endswith(('.xls', '.xlsx', '.xlsb'))]
            if not arquivos:
                raise FileNotFoundError("Nenhuma planilha encontrada!\n"
                "Coloque as planilhas na\n"
                "pasta Extras Planilhas")
                
    
            self.caminhos_comparacao = arquivos
            self.label_comparacao.config(
                text=f"{len(arquivos)} Planilhas carregadas!",
                fg="white",
                font=("Arial", 10, "bold")
            )

            # Indicação visual de sucesso
            self.botao_comparacao.config(bg="#4CAF50")  # Verde
            self.root.after(5000, lambda: self.reset_cor_botao(self.botao_comparacao))  
        except Exception as e:
            self.label_comparacao.config(text=f"Erro: {e}", fg="#ff5a5a")
            self.botao_comparacao.config(bg="#fc6a6a")  # Vermelho



    def comparar_planilhas(self):
        diretorio_processado = r"C:/Comparador de Planilhas/Extras Planilhas Processadas/"
        try:
            if self.caminho_mestre and self.caminhos_comparacao:
                df_resultado = comparar_planilhas(self.caminho_mestre, self.caminhos_comparacao)

                caminho_saida = gerar_nome_arquivo_sugerido()
                os.makedirs(os.path.dirname(caminho_saida), exist_ok=True)

                salvar_planilha_com_estilo(df_resultado, caminho_saida)

                import shutil
                os.makedirs(diretorio_processado, exist_ok=True)

                for caminho in self.caminhos_comparacao:
                    nome_arquivo = os.path.basename(caminho)
                    destino = os.path.join(diretorio_processado, nome_arquivo)

                    contador = 1
                    base, extensao = os.path.splitext(nome_arquivo)
                    while os.path.exists(destino):
                        destino = os.path.join(diretorio_processado, f"{base}_{contador:02d}{extensao}")
                        contador += 1

                    shutil.move(caminho, destino)

                # Indicação visual de sucesso
                self.botao_comparar.hover_ativo = False  # desativa o hover
                self.botao_comparar.config(bg="#4CAF50")  # Verde
                self.label_movido.config(
                    text="Planilhas movidas para 'Extras Planilhas Processadas'.",
                    fg="#00ff00"
                )


                # Garante cancelamento prévio caso já exista timer
                if self.botao_comparar.after_id:
                    self.root.after_cancel(self.botao_comparar.after_id)

                self.root.after(5000, lambda: self.reset_cor_botao(self.botao_comparar))

                self.caminhos_comparacao = []
                self.label_comparacao.config(text="Nenhuma planilha carregada.")

            else:
                raise ValueError("Planilha mestre ou planilhas de \n"
                "comparação não foram selecionadas!")

        except Exception as e:
            self.botao_comparar.config(bg="#ff5a5a")  # Vermelho
            self.label_movido.config(
                text=f"Erro: {e}",
                fg="#ff4444"
            )

            self.root.after(3000, lambda: self.reset_cor_botao(self.botao_comparar))



    def abrir_pasta_viagens(self):
        pasta_destino = r"C:/Comparador de Planilhas/Viagens do dia/"
        try:
            os.makedirs(pasta_destino, exist_ok=True)
            os.startfile(pasta_destino)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a pasta: {e}")



    def reset_cor_botao(self, botao):
        botao.config(bg="#1c294a")
        botao.hover_ativo = True
        botao.after_id = None  # limpa o ID após execução

    

         
                
# ===============================
# MAIN
# ===============================
if __name__ == "__main__":
    root = tk.Tk()
    app = ComparadorPlanilhas(root)
    root.mainloop()